import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.models import Species

ALLOWED_CLASSIFICATIONS = {"CR", "EN", "VU"}


def clean_value(value):
    if value is None:
        return None
    text = str(value).strip()
    if text.upper() in {"", "N/A", "NA", "NONE"}:
        return None
    return text


class Command(BaseCommand):
    help = "Import threatened species from the cleaned Excel dataset."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Optional sheet name. If omitted, the active sheet is used.",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        sheet_name = options["sheet"]

        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {excel_path}")

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Sheet '{sheet_name}' not found in workbook.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise CommandError("The selected sheet is empty.")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

        required_columns = {
            "source_sheet",
            "common_name",
            "scientific_name",
            "nt_classification",
        }

        missing = required_columns - set(headers)
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(sorted(missing))}")

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in rows[1:]:
            row_data = dict(zip(headers, row))

            nt_classification = clean_value(row_data.get("nt_classification"))
            if nt_classification not in ALLOWED_CLASSIFICATIONS:
                skipped_count += 1
                continue

            scientific_name = clean_value(row_data.get("scientific_name"))
            if not scientific_name:
                skipped_count += 1
                continue

            defaults = {
                "category": clean_value(row_data.get("source_sheet")),
                "common_name": clean_value(row_data.get("common_name")) or scientific_name,
                "nt_classification": nt_classification,
                "introduced_status": clean_value(row_data.get("introduced_status")),
            }

            # Optional fields if they exist in your file
            if "family" in headers:
                defaults["family"] = clean_value(row_data.get("family"))

            if "order_name" in headers:
                defaults["order_name"] = clean_value(row_data.get("order_name"))

            if "epbc_classification" in headers:
                defaults["epbc_classification"] = clean_value(row_data.get("epbc_classification"))

            obj, created = Species.objects.update_or_create(
                scientific_name=scientific_name,
                defaults=defaults,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Created: {created_count}, "
                f"Updated: {updated_count}, Skipped: {skipped_count}"
            )
        )